import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
import os
import sys

#scrapes data from a journal sheet and creates a ledger from said data
#i made this because i was tired of copying and pasting
#theres probably already a simpler way to do this
#but hey, python

class Journal:
    def __init__(self):
        self.company_name = ""
        self.date_range = ""
        self.transactions = []

    def print_info(self):
        print(self.company_name)
        print(self.date_range)

        #print every transaction nicely
        for transaction in self.transactions:
            print()
            print('Date: ' + str(transaction.date))
            #print all debits
            for i in range(0,len(transaction.debits)):
                print('Debit ' + transaction.debits[i].account_name + ' for ' + str(transaction.debits[i].value))
            #print all credits
            for i in range(0,len(transaction.credits)):
                print('Credit ' + transaction.credits[i].account_name + ' for ' + str(transaction.credits[i].value))

            #print description
            print(transaction.description)


    def scrape_journal_sheet(self, ws):
        self.company_name = worksheet['A1'].value
        self.date_range = worksheet['A3'].value

        dateCellsIndeces = []
        endPointIndeces = []
        debitEndIndeces = []
        descriptionIndeces = []

        #find cell of each transaction's date stamp and record it
        for i in range(6,151): #arbitrary upper bound, will automate when i figure out how
            if ws['A'+str(i)].value:
                dateCellsIndeces.append(i)

        #find endpoints of every individual transactions
        #start by getting endpoints for every transaction besides the last one by subtracting one from the date index array
        for entry in dateCellsIndeces:
            #if it is not the first transaction
            if entry != 6:
                endPointIndeces.append(entry - 1)
        endPointIndeces.append(dateCellsIndeces[len(dateCellsIndeces)-1]+3) #hardcoded + 3 assuming its an adjusting entry, will automate when i figure out how

        #loop through each transaction by the date cell index, record debits and credits
        #and put them in lists of entries to be put into a transaction, then added
        #to the transactions list of this class

        for date_iterator in range(0,len(dateCellsIndeces)):
            #find end of debits (where a cell is blank and, because of double entry accounting, a credit must occur)
            debit_end_index = 1
            for i in range(dateCellsIndeces[date_iterator],endPointIndeces[date_iterator]):
                if not ws['B'+str(i)].value:
                    debitEndIndeces.append(i)
                    break #maybe not best practice but hey it works

        #find the description cell by just using the one after the end

        for description_iterator in range(0,len(endPointIndeces)):
            for i in range(debitEndIndeces[description_iterator],endPointIndeces[description_iterator]):
                if ws['B'+str(i)].value:
                    descriptionIndeces.append(i)

        '''
        print('date row indeces')
        print(str(dateCellsIndeces))
        print('end of transaction row indeces')
        print(str(endPointIndeces))
        print('end of debit (start of credit) row indeces')
        print(str(debitEndIndeces))
        print('description row indeces')
        print(str(descriptionIndeces))
        '''

        #record each transaction
        for transaction_iterator in range(0,len(dateCellsIndeces)):
            #record all debit entries
            debit_entries = []
            for i in range(dateCellsIndeces[transaction_iterator],debitEndIndeces[transaction_iterator]):
                new_debit_entry = Entry(ws['B'+str(i)].value,ws['D'+str(i)].value, True)
                debit_entries.append(new_debit_entry)

            #record all credit entries
            credit_entries = []
            for i in range(debitEndIndeces[transaction_iterator],descriptionIndeces[transaction_iterator]):
                new_credit_entry = Entry(ws['C'+str(i)].value,ws['E'+str(i)].value, False)
                credit_entries.append(new_credit_entry)

            #MAKE A SPICY TRANSACTION OBJECT AND PUT IT INTO THE CLASS TRANSACTION LIST
            new_transaction = Transaction(ws['A'+str(dateCellsIndeces[transaction_iterator])].value,debit_entries,credit_entries,ws['B'+str(descriptionIndeces[transaction_iterator])].value)
            self.transactions.append(new_transaction)

#represents one line in a journal
#ie: a debit of x account for y value
class Entry:
    def __init__(self, account_name, value, is_debit):
        self.account_name = account_name
        self.value = value
        self.is_debit = is_debit
        self.date = ""


class Transaction:
    def __init__(self, date, debits, credits, description):
        self.date = date
        self.debits = debits
        self.credits = credits
        self.description = description

class Account:
    def __init__(self, account_name):
        self.account_name = account_name
        self.entries = []

class Ledger:
    def __init__(self):
        self.company_name = ""
        self.date_range = ""
        self.accounts = []
        self.output_excel_filename = ""

    #returns a tuple with a bool if account exists and an int of the index in the class accounts list
    def does_account_exist(self, account_name):
        (does_exist,index) = (False,0)
        for account_index in range(len(self.accounts)):
            if account_name == self.accounts[account_index].account_name:
                (does_exist,index) = (True, account_index)
        return (does_exist,index)

    #takes a journal object and creates ledger data
    def import_journal(self, journal):
        #get identifying data from journal
        self.company_name = journal.company_name
        self.date_range = journal.date_range

        #go through each transaction
        for transaction_index in range(len(journal.transactions)):
            #get date
            transaction_date = journal.transactions[transaction_index].date

            #iterate through debits first
            for debit_index in range(len(journal.transactions[transaction_index].debits)):
                #assign debit the date previously gotten
                journal.transactions[transaction_index].debits[debit_index].date = transaction_date

                (account_exists, account_index) = self.does_account_exist(journal.transactions[transaction_index].debits[debit_index].account_name)
                #if the account already exists, just append it to the entry list
                if account_exists:
                    self.accounts[account_index].entries.append(journal.transactions[transaction_index].debits[debit_index])
                else: #if it doesn't, create a new account with that name and then append it
                    new_account = Account(journal.transactions[transaction_index].debits[debit_index].account_name)
                    self.accounts.append(new_account)
                    self.accounts[len(self.accounts)-1].entries.append(journal.transactions[transaction_index].debits[debit_index])

            #iterate through credits now
            for credit_index in range(len(journal.transactions[transaction_index].credits)):
                #assign debit the date previously gotten
                journal.transactions[transaction_index].credits[credit_index].date = transaction_date

                #*
                (account_exists, account_index) = self.does_account_exist(journal.transactions[transaction_index].credits[credit_index].account_name)
                #if the account already exists, just append it to the entry list
                if account_exists:
                    self.accounts[account_index].entries.append(journal.transactions[transaction_index].credits[credit_index])
                else: #if it doesn't, create a new account with that name and then append it
                    new_account = Account(journal.transactions[transaction_index].credits[credit_index].account_name)
                    self.accounts.append(new_account)
                    self.accounts[len(self.accounts)-1].entries.append(journal.transactions[transaction_index].credits[credit_index])

    def print_ledger(self):

        print(self.company_name)
        print("General Ledger")
        print(self.date_range)

        #print each ledger account
        for account in self.accounts:
            print()
            print(str(account.account_name))
            for entry in account.entries:
                word = "Debit"
                if entry.is_debit == False:
                    word = "Credit"
                print(str(entry.date) + ' : ' + word + ' ' + str(entry.value))

    def ledger_worksheet(self):

        new_workbook = Workbook() #temporary
        worksheet = new_workbook.active

        worksheet['A1'] = self.company_name
        worksheet['A2'] = "General Ledger"
        worksheet['A3'] = self.date_range

        worksheet.column_dimensions['A'].width = 7
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 10

        worksheet.merge_cells('A1:E1')
        worksheet.merge_cells('A2:E2')
        worksheet.merge_cells('A3:E3')


        #first account starts at row 5
        #get indeces of when to start accounts
        account_start_indeces = []
        account_start_indeces.append(5)
        for account_index in range(len(self.accounts)-1): #make sure to not iterate through the last account in the ledger
            #print(account_index)
            #print(len(self.accounts))
            number = account_start_indeces[account_index] + 1
            number += len(self.accounts[account_index].entries) + 3
            account_start_indeces.append(number)
            set_border(worksheet,'A'+str(account_start_indeces[account_index])+':E'+str(number-3),'medium')
            set_border(worksheet,'A'+str(account_start_indeces[account_index])+':E'+str(account_start_indeces[account_index]),'medium')


        for start_index in range(len(account_start_indeces)):
            #put labels where they need to go
            worksheet['A'+str(account_start_indeces[start_index])] = self.accounts[start_index].account_name
            #merge name cell
            worksheet.merge_cells('A'+str(account_start_indeces[start_index])+':E'+str(account_start_indeces[start_index]))

            worksheet['A'+str(account_start_indeces[start_index]+1)] = "Date"
            worksheet['B'+str(account_start_indeces[start_index]+1)] = "Description"
            worksheet['C'+str(account_start_indeces[start_index]+1)] = "Debit"
            worksheet['D'+str(account_start_indeces[start_index]+1)] = "Credit"
            worksheet['E'+str(account_start_indeces[start_index]+1)] = "Balance"


            #log entries
            for entry_index in range(len(self.accounts[start_index].entries)): #iterate through every entry
                row_number = account_start_indeces[start_index] + entry_index + 2

                #value in correct column
                column_string = 'C'
                if not self.accounts[start_index].entries[entry_index].is_debit:
                    column_string = 'D'
                worksheet[column_string+str(row_number)] = self.accounts[start_index].entries[entry_index].value

                #date
                worksheet['A'+str(row_number)] = self.accounts[start_index].entries[entry_index].date
                #format the cell
                worksheet['A'+str(row_number)].number_format = 'MM/DD'


        new_workbook.save(self.output_excel_filename)

#https://stackoverflow.com/questions/34520764/apply-border-to-range-of-cells-using-openpyxl (modified slightly)
def set_border(ws, cell_range, border_style):
    rows = ws[cell_range]
    side = Side(border_style=border_style, color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

def main(input_worksheet):

    journal = Journal()
    journal.scrape_journal_sheet(input_worksheet)
    #journal.print_info()

    ledger = Ledger()
    ledger.import_journal(journal)
    ledger.output_excel_filename = str(output_excel_filename)
    #ledger.print_ledger()
    ledger.ledger_worksheet()


#CODE PAST THIS RUNS AT START------------------------------------------------------------------------

'''
    Argument Format in terminal

    python3 autoledger.py (input file name) (number of sheet that the ledger u want is on) (output file name)

    ie:

    python3 autoledger.py journal.xlsx 1 output.xlsx
'''

sytem_argv = sys.argv
input_excel_file = sys.argv[1]
input_sheet_number = int(sys.argv[2])
output_excel_filename = sys.argv[3]

#print(sys.argv)
print('Input: ' + input_excel_file + ' Sheet ' + str(input_sheet_number))
print('Output: ' + output_excel_filename)


#this assumes that the journal spreadsheet is formatted exactly correct, otherwise everything will probably break because it is so specific

workbook = load_workbook(input_excel_file, data_only = True,read_only = True)
worksheet = workbook.worksheets[input_sheet_number-1] #-1 because indeces and humans dont work the same way

main(worksheet)

print("Success! (hopefully)")
os.system('open '+output_excel_filename)
